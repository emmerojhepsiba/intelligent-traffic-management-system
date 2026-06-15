import os
import uuid
import cv2
import numpy as np
import threading
import time
from datetime import datetime
from flask import Blueprint, request, jsonify, render_template
from routes.auth import login_required
from utils.db import get_db
from utils.detection import detect_vehicles, draw_detections
from config import Config

dashboard_bp = Blueprint("dashboard", __name__)


def _safe_result(result: dict) -> dict:
    """Strip numpy types so the dict is JSON/MongoDB safe."""
    clean = {}
    for k, v in result.items():
        if isinstance(v, (np.integer,)):
            clean[k] = int(v)
        elif isinstance(v, (np.floating,)):
            clean[k] = float(v)
        elif isinstance(v, np.ndarray):
            clean[k] = v.tolist()
        elif isinstance(v, list):
            clean[k] = [
                {kk: (int(vv) if isinstance(vv, np.integer) else
                      float(vv) if isinstance(vv, np.floating) else vv)
                 for kk, vv in item.items()} if isinstance(item, dict) else item
                for item in v
            ]
        elif isinstance(v, dict):
            clean[k] = {kk: (int(vv) if isinstance(vv, np.integer) else
                              float(vv) if isinstance(vv, np.floating) else vv)
                        for kk, vv in v.items()}
        else:
            clean[k] = v
    return clean


@dashboard_bp.route("/dashboard")
@login_required
def dashboard_page():
    return render_template("dashboard.html", admin=request.admin)


@dashboard_bp.route("/api/dashboard/stats")
@login_required
def get_stats():
    try:
        db = get_db()
        total_detections = db.detections.count_documents({})
        total_incidents  = db.incidents.count_documents({})
        active_incidents = db.incidents.count_documents({"status": "active"})
        emergency_count  = db.detections.count_documents({"emergency_detected": True})
        latest = db.detections.find_one(sort=[("timestamp", -1)])
        current_density  = latest["density"] if latest else "low"
        return jsonify({
            "total_detections": total_detections,
            "total_incidents":  total_incidents,
            "active_incidents": active_incidents,
            "emergency_count":  emergency_count,
            "current_density":  current_density,
        })
    except Exception as e:
        print(f"[Dashboard] Stats error: {e}")
        return jsonify({
            "total_detections": 0, "total_incidents": 0,
            "active_incidents": 0, "emergency_count": 0,
            "current_density": "low",
        })


@dashboard_bp.route("/api/dashboard/upload", methods=["POST"])
@login_required
def upload_media():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    location = request.form.get("location", "Unknown Location")

    if not file or file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    is_image = ext in Config.ALLOWED_IMAGE_EXT
    is_video = ext in Config.ALLOWED_VIDEO_EXT

    if not (is_image or is_video):
        return jsonify({"error": f"File type .{ext} not supported"}), 400

    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(Config.UPLOAD_FOLDER, filename)

    try:
        file.save(filepath)
    except Exception as e:
        return jsonify({"error": f"Could not save file: {e}"}), 500

    # ── Image processing ────────────────────────────────────────────────────
    if is_image:
        try:
            image = cv2.imread(filepath)
            if image is None:
                return jsonify({"error": "Could not read image — file may be corrupt"}), 400

            result = detect_vehicles(image_array=image)
            annotated = draw_detections(image, result["detections"])

            annotated_filename = f"det_{filename}"
            annotated_path = os.path.join(Config.UPLOAD_FOLDER, annotated_filename)
            cv2.imwrite(annotated_path, annotated)

            result = _safe_result(result)
            result["media_type"]     = "image"
            result["original_file"]  = f"/static/uploads/{filename}"
            result["annotated_file"] = f"/static/uploads/{annotated_filename}"
            result["location"]       = location

            try:
                db = get_db()
                db.detections.insert_one({**result, "timestamp": datetime.utcnow()})
            except Exception as e:
                print(f"[Dashboard] DB insert error: {e}")

            if result.get("emergency_detected"):
                try:
                    _create_emergency_alert(result, location)
                    _trigger_emergency_signals(result, location)
                except Exception as e:
                    print(f"[Dashboard] Emergency alert error: {e}")

            return jsonify(result)

        except Exception as e:
            print(f"[Dashboard] Image processing error: {e}")
            return jsonify({"error": f"Processing failed: {str(e)}"}), 500

    # ── Video processing ────────────────────────────────────────────────────
    elif is_video:
        try:
            cap = cv2.VideoCapture(filepath)
            if not cap.isOpened():
                return jsonify({"error": "Could not open video — file may be corrupt or unsupported"}), 400

            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) or 1
            fps          = cap.get(cv2.CAP_PROP_FPS) or 25
            duration_s   = total_frames / fps

            NUM_SAMPLES = min(16, max(1, total_frames))
            sample_positions = [int(i * total_frames / NUM_SAMPLES) for i in range(NUM_SAMPLES)]

            frame_results = []

            for pos in sample_positions:
                try:
                    cap.set(cv2.CAP_PROP_POS_FRAMES, pos)
                    ret, frame = cap.read()
                    if not ret or frame is None:
                        continue
                    r = detect_vehicles(image_array=frame)
                    frame_results.append((r, frame.copy()))
                except Exception as fe:
                    print(f"[Dashboard] Frame {pos} error: {fe}")
                    continue

            cap.release()

            if not frame_results:
                return jsonify({"error": "Could not extract any frames from video"}), 400

            # Aggregate across all frames
            # Use mode (most common count) to avoid inflating count from noisy frames
            from collections import Counter
            emerg_counts = [r["emergency_count"] for r, _ in frame_results]
            count_mode = Counter(emerg_counts).most_common(1)[0][0]
            # Cap at mode — never report more than the most-seen count across frames
            final_emerg_count = count_mode

            all_emerg_types = list({
                etype
                for r, _ in frame_results
                for etype in r.get("emergency_vehicles", [])
            })

            # Pick the best frame: one whose emergency_count matches the mode,
            # then highest total vehicles for a representative annotated image
            mode_frames = [(r, f) for r, f in frame_results if r["emergency_count"] == count_mode]
            best_result, best_frame = max(
                mode_frames if mode_frames else frame_results,
                key=lambda x: x[0]["total_vehicles"]
            )

            # Deep-copy best_result so we don't mutate the original
            best_result = dict(best_result)
            best_result["emergency_count"]    = int(final_emerg_count)
            best_result["emergency_vehicles"] = all_emerg_types
            best_result["emergency_detected"] = final_emerg_count > 0
            best_result["video_duration_s"]   = round(duration_s, 1)
            best_result["video_frames_sampled"] = int(NUM_SAMPLES)

            annotated = draw_detections(best_frame, best_result["detections"])
            annotated_filename = f"det_{uuid.uuid4().hex}.jpg"
            annotated_path = os.path.join(Config.UPLOAD_FOLDER, annotated_filename)
            cv2.imwrite(annotated_path, annotated)

            best_result = _safe_result(best_result)
            best_result["media_type"]     = "video"
            best_result["original_file"]  = f"/static/uploads/{filename}"
            best_result["annotated_file"] = f"/static/uploads/{annotated_filename}"
            best_result["location"]       = location

            try:
                db = get_db()
                db.detections.insert_one({**best_result, "timestamp": datetime.utcnow()})
            except Exception as e:
                print(f"[Dashboard] DB insert error: {e}")

            if best_result.get("emergency_detected"):
                try:
                    _create_emergency_alert(best_result, location)
                    _trigger_emergency_signals(best_result, location)
                except Exception as e:
                    print(f"[Dashboard] Emergency alert error: {e}")

            return jsonify(best_result)

        except Exception as e:
            print(f"[Dashboard] Video processing error: {e}")
            return jsonify({"error": f"Video processing failed: {str(e)}"}), 500


@dashboard_bp.route("/api/dashboard/detections")
@login_required
def get_detections():
    try:
        db = get_db()
        detections = list(
            db.detections.find({}, {"_id": 0})
            .sort("timestamp", -1)
            .limit(50)
        )
        return jsonify(detections)
    except Exception as e:
        print(f"[Dashboard] Get detections error: {e}")
        return jsonify([])


@dashboard_bp.route("/api/dashboard/incidents")
@login_required
def get_incidents():
    try:
        db = get_db()
        incidents = list(
            db.incidents.find({}, {"_id": 0})
            .sort("timestamp", -1)
            .limit(50)
        )
        return jsonify(incidents)
    except Exception as e:
        print(f"[Dashboard] Get incidents error: {e}")
        return jsonify([])


@dashboard_bp.route("/api/dashboard/incidents", methods=["POST"])
@login_required
def create_incident():
    try:
        data = request.get_json() or {}
        db = get_db()
        incident = {
            "id": uuid.uuid4().hex[:12],
            "type": data.get("type", "general"),
            "location": data.get("location", ""),
            "description": data.get("description", ""),
            "severity": data.get("severity", "medium"),
            "status": "active",
            "timestamp": datetime.utcnow(),
            "reported_by": request.admin.get("email", "unknown"),
        }
        db.incidents.insert_one(incident)
        incident.pop("_id", None)
        return jsonify(incident), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@dashboard_bp.route("/api/dashboard/incidents/<incident_id>/resolve", methods=["POST"])
@login_required
def resolve_incident(incident_id):
    try:
        db = get_db()
        result = db.incidents.update_one(
            {"id": incident_id},
            {"$set": {"status": "resolved", "resolved_at": datetime.utcnow()}}
        )
        if result.modified_count:
            return jsonify({"message": "Incident resolved"})
        return jsonify({"error": "Incident not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@dashboard_bp.route("/api/dashboard/signals", methods=["GET"])
@login_required
def get_signals():
    """Return the latest state for every signal line."""
    try:
        db = get_db()
        states = {}
        for lid in SIGNAL_LINES:
            latest = db.traffic_logs.find_one(
                {"type": "signal_change", "line_id": lid},
                sort=[("timestamp", -1)]
            )
            states[lid] = latest["state"] if latest else "red"
        return jsonify(states)
    except Exception as e:
        print(f"[Signal] get_signals error: {e}")
        return jsonify({lid: "red" for lid in SIGNAL_LINES})


@dashboard_bp.route("/api/dashboard/signals", methods=["POST"])
@login_required
def update_signal():
    try:
        data = request.get_json() or {}
        line_id     = data.get("line_id")
        signal_state = data.get("state")

        if signal_state not in ("green", "yellow", "red"):
            return jsonify({"error": "Invalid signal state"}), 400

        db = get_db()
        db.traffic_logs.insert_one({
            "type": "signal_change",
            "line_id": line_id,
            "state": signal_state,
            "timestamp": datetime.utcnow(),
            "changed_by": request.admin.get("email", "unknown"),
        })
        return jsonify({
            "message": f"Signal for {line_id} changed to {signal_state}",
            "line_id": line_id,
            "state": signal_state,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@dashboard_bp.route("/api/dashboard/signals/emergency", methods=["POST"])
@login_required
def emergency_signal():
    """
    On emergency detection: set the detected line to GREEN for 60s,
    all other lines to RED. After 60s, revert all lines to normal cycle.
    """
    try:
        data = request.get_json() or {}
        priority_line = data.get("line_id")
        all_lines     = data.get("all_lines", [])
        triggered_by  = request.admin.get("email", "system")

        if not priority_line:
            return jsonify({"error": "line_id required"}), 400

        # Import socketio lazily to avoid circular import
        from app import socketio

        db = get_db()

        def _emit_and_log(line_id, state, changed_by):
            try:
                db.traffic_logs.insert_one({
                    "type": "signal_change",
                    "line_id": line_id,
                    "state": state,
                    "timestamp": datetime.utcnow(),
                    "changed_by": changed_by,
                    "reason": "emergency_vehicle" if line_id == priority_line else "emergency_clearance",
                })
                socketio.emit("signal_changed", {
                    "line_id": line_id,
                    "state": state,
                    "emergency": True,
                    "priority_line": priority_line,
                })
            except Exception as e:
                print(f"[Signal] emit/log error: {e}")

        def _run_sequence():
            # Phase 1: priority line GREEN, all others RED
            for lid in all_lines:
                state = "green" if lid == priority_line else "red"
                _emit_and_log(lid, state, triggered_by)

            # Hold GREEN for 60s
            time.sleep(60)

            # Phase 2: priority line YELLOW (transition)
            _emit_and_log(priority_line, "yellow", "system")
            time.sleep(5)

            # Phase 3: revert — priority line RED, others back to green
            for lid in all_lines:
                state = "red" if lid == priority_line else "green"
                _emit_and_log(lid, state, "system")

        # Run in background so the HTTP response returns immediately
        t = threading.Thread(target=_run_sequence, daemon=True)
        t.start()

        return jsonify({
            "message": f"Emergency signal sequence started for {priority_line}",
            "priority_line": priority_line,
            "green_duration_s": 60,
            "other_lines_state": "red",
        })

    except Exception as e:
        print(f"[Signal] Emergency signal error: {e}")
        return jsonify({"error": str(e)}), 500


@dashboard_bp.route("/api/dashboard/export/<report_type>/<fmt>")
@login_required
def export_report(report_type, fmt):
    from flask import Response
    db = get_db()

    # ── Fetch data ────────────────────────────────────────────────────────
    if report_type == "detections":
        headers = ["Timestamp", "Location", "Total Vehicles", "Density",
                   "Emergency Detected", "Emergency Count", "Emergency Vehicles", "Media Type"]
        rows = []
        for d in db.detections.find({}, {"_id": 0}).sort("timestamp", -1).limit(500):
            rows.append([
                str(d.get("timestamp", "")), d.get("location", ""),
                d.get("total_vehicles", 0), d.get("density", ""),
                d.get("emergency_detected", False), d.get("emergency_count", 0),
                ", ".join(d.get("emergency_vehicles", [])), d.get("media_type", "image")
            ])
    elif report_type == "incidents":
        headers = ["Timestamp", "ID", "Type", "Location", "Description",
                   "Severity", "Status", "Reported By", "Resolved At"]
        rows = []
        for d in db.incidents.find({}, {"_id": 0}).sort("timestamp", -1).limit(500):
            rows.append([
                str(d.get("timestamp", "")), d.get("id", ""), d.get("type", ""),
                d.get("location", ""), d.get("description", ""),
                d.get("severity", ""), d.get("status", ""),
                d.get("reported_by", ""), str(d.get("resolved_at", ""))
            ])
    elif report_type == "signals":
        headers = ["Timestamp", "Line ID", "State", "Reason", "Changed By"]
        rows = []
        for d in db.traffic_logs.find({"type": "signal_change"}, {"_id": 0}).sort("timestamp", -1).limit(500):
            rows.append([
                str(d.get("timestamp", "")), d.get("line_id", ""),
                d.get("state", ""), d.get("reason", ""), d.get("changed_by", "")
            ])
    else:
        return jsonify({"error": "Invalid report type"}), 400

    title = {"detections": "Detection Report", "incidents": "Incidents Report", "signals": "Signal Logs Report"}[report_type]

    # ── Excel export ──────────────────────────────────────────────────────
    if fmt == "excel":
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = title

        # Title row
        ws.merge_cells(f"A1:{chr(64 + len(headers))}1")
        ws["A1"] = f"Smart Traffic Management — {title}"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1e3a5f")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Header row
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2d5a9e")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r, row in enumerate(rows, 3):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={report_type}_report.xlsx"}
        )

    # ── PDF export ────────────────────────────────────────────────────────
    elif fmt == "pdf":
        import io
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=1*cm, rightMargin=1*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(f"Smart Traffic Management — {title}", styles["Title"]))
        elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Total records: {len(rows)}", styles["Normal"]))
        elements.append(Spacer(1, 0.4*cm))

        # Table
        data = [headers] + [[str(v) for v in row] for row in rows]
        col_count = len(headers)
        col_width = (landscape(A4)[0] - 2*cm) / col_count
        t = Table(data, colWidths=[col_width] * col_count, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0),  8),
            ("FONTSIZE",    (0, 1), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4ff")]),
            ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        doc.build(elements)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={report_type}_report.pdf"}
        )

    return jsonify({"error": "Invalid format"}), 400


@dashboard_bp.route("/api/dashboard/logs")
@login_required
def get_logs():
    try:
        db = get_db()
        logs = list(
            db.traffic_logs.find({}, {"_id": 0})
            .sort("timestamp", -1)
            .limit(100)
        )
        return jsonify(logs)
    except Exception as e:
        print(f"[Dashboard] Get logs error: {e}")
        return jsonify([])


def _create_emergency_alert(detection_result, location):
    db = get_db()
    vehicles = detection_result.get("emergency_vehicles") or ["emergency vehicle"]
    alert = {
        "id": uuid.uuid4().hex[:12],
        "type": "emergency_vehicle",
        "location": location,
        "description": f"Emergency vehicle(s) detected: {', '.join(vehicles)}",
        "severity": "critical",
        "status": "active",
        "emergency_vehicles": vehicles,
        "timestamp": datetime.utcnow(),
        "reported_by": "AI Detection System",
    }
    db.incidents.insert_one(alert)


# All configured signal lines — must match the JS signalLines array
SIGNAL_LINES = ["line-1", "line-2", "line-3", "line-4", "line-5", "line-6"]


def _trigger_emergency_signals(detection_result, location):
    """
    Parse the line id from the location string and kick off the
    emergency signal sequence in a background thread.
    Location format expected: "Line 1 - ..." or "line-1" etc.
    """
    try:
        from app import socketio, create_app

        priority_line = _parse_line_id(location)
        if not priority_line:
            print(f"[Signal] Could not parse line id from location: {location!r}")
            return

        # Grab the current Flask app so we can push its context into the thread
        from flask import current_app
        app = current_app._get_current_object()

        def _emit_and_log(line_id, state, reason="emergency_clearance"):
            try:
                with app.app_context():
                    db = get_db()
                    db.traffic_logs.insert_one({
                        "type": "signal_change",
                        "line_id": line_id,
                        "state": state,
                        "timestamp": datetime.utcnow(),
                        "changed_by": "AI Detection System",
                        "reason": reason,
                    })
                socketio.emit("signal_changed", {
                    "line_id": line_id,
                    "state": state,
                    "emergency": True,
                    "priority_line": priority_line,
                }, namespace="/")
                print(f"[Signal] Emitted {line_id} → {state}")
            except Exception as e:
                print(f"[Signal] emit/log error: {e}")

        def _run_sequence():
            print(f"[Signal] Sequence START — {priority_line} GREEN, others RED")
            # Phase 1 — priority line GREEN (60s), all others RED
            for lid in SIGNAL_LINES:
                reason = "emergency_vehicle" if lid == priority_line else "emergency_clearance"
                _emit_and_log(lid, "green" if lid == priority_line else "red", reason)

            time.sleep(60)

            # Phase 2 — priority line YELLOW (5s transition)
            _emit_and_log(priority_line, "yellow", "emergency_end")
            time.sleep(5)

            # Phase 3 — revert: priority line RED, others GREEN
            print(f"[Signal] Sequence END — reverting all lines")
            for lid in SIGNAL_LINES:
                _emit_and_log(lid, "red" if lid == priority_line else "green", "emergency_end")

        threading.Thread(target=_run_sequence, daemon=True).start()
        print(f"[Signal] Emergency sequence started — {priority_line} GREEN for 60s")

    except Exception as e:
        print(f"[Signal] _trigger_emergency_signals error: {e}")


def _parse_line_id(location: str) -> str:
    """
    Extract a line id from a free-text location string.
    Handles: 'Line 1 - Mangamur Road', 'line-3', 'LINE 4', 'line 2', etc.
    Returns e.g. 'line-1' or None if not found.
    """
    import re
    if not location:
        return None
    m = re.search(r'line[-\s]?(\d)', location, re.IGNORECASE)
    if m:
        return f"line-{m.group(1)}"
    return None
